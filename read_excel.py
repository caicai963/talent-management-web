import openpyxl, re

path = r'C:\Users\wb.liujing23\AppData\Local\Temp\lobsterai\attachments\(UX)研究资源成本估算工具-v2.6-1775099052261-gae2yf.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

print('Sheet names:', wb.sheetnames)
print()

# Read the 国内研究成本估算 sheet
if '国内研究成本估算' in wb.sheetnames:
    ws = wb['国内研究成本估算']
    print('Sheet: 国内研究成本估算')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    print()
    
    # Show C36-C39 area (rows 34-42, cols A-J)
    print('=== Rows 34-42, Cols A-J ===')
    for row in range(34, 43):
        row_data = []
        for col in range(1, 11):
            cell = ws.cell(row, col)
            row_data.append(repr(cell.value)[:30])
        print(f'Row {row}: ' + ' | '.join(row_data))
    
    print()
    # Also show col C and H specifically
    print('=== Col C (34-42) and Col H (34-42) ===')
    for row in range(34, 43):
        c_val = ws.cell(row, 3).value  # Col C
        h_val = ws.cell(row, 8).value  # Col H
        print(f'Row {row}: C={repr(c_val)[:40]}, H={repr(h_val)[:40]}')
